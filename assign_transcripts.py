# %%
import random

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# %%

from moves.library import start

# %%
files = os.listdir(start.SHARED_PATH + "excel transcripts/")


# Create random order of files for coding
random.seed(10)
random.shuffle(files)
random_df = pd.DataFrame(files)
random_df.to_csv(start.SHARED_PATH + "assignments.csv")

random_order = pd.read_csv(start.SHARED_PATH + "assignments.csv")

transcript_files = list(random_order["0"])
transcript_files = [f[:-4] + "xlsx" for f in transcript_files]

# %%

df = pd.DataFrame()
for filename in transcript_files:
    file_path = start.SHARED_PATH + "excel transcripts/"
    data = pd.read_excel(file_path + filename)
    data["doc"] = filename
    df = df.append(data)


df["turn_count"] = df.groupby("doc").cumcount()
df = df[["doc", "Time-stamp", "turn_count", "Speaker", "Text"]]

coach_df = df[df.Speaker == "Coach"]
teacher_df = df[df.Speaker != "Coach"]
teacher_df["turn_count"] = teacher_df.turn_count + 1

teacher_df = teacher_df.rename(columns={"Text": "preceding_teacher_text"})

new_df = coach_df.merge(
    teacher_df[["preceding_teacher_text", "doc", "turn_count"]],
    how="left",
    left_on=["doc", "turn_count"],
    right_on=["doc", "turn_count"],
)
new_df = new_df[["doc", "turn_count", "preceding_teacher_text", "Text"]]

# %%
new_df["id"] = new_df.reset_index().index

new_df = new_df[["id", "doc", "turn_count", "preceding_teacher_text", "Text"]]

new_df.to_csv(start.SHARED_PATH + "utterance_id.csv", index=False)

new_df = pd.read_csv(start.SHARED_PATH + "utterance_id.csv")

# %% Select 10 transcripts for coding
pilot1_list = transcript_files[0:10]

pilot1 = new_df[new_df.doc.isin(pilot1_list)]

# %% Out-of-Context
move = "TellBack Positive Evaluation"
wb = Workbook()
ws = wb.active
ws["B1"] = "TIME SPENT CODING IN MINUTES (B1) TO THE NEAREST SECOND (C1) "
ws["B2"] = "MOVE: "
ws["C2"] = move
ws["C3"] = "Text"
ws["C3"] = "Code"

pilot1_out_of_context = pilot1.sample(len(pilot1), random_state=10).rename(
    columns={"preceding_teacher_text": "Preceding Teacher Text"}
)
for r in dataframe_to_rows(
    pilot1_out_of_context[["id", "Text"]], index=False, header=False
):
    ws.append(r)

wb.save(start.SHARED_PATH + "coding files/Pilot 1 Out-of-Context/1 " + move + ".xlsx")

# %%
