# %%
import random

import pandas as pd
from openpyxl import load_workbook


from moves.library import start


RANDOM_SEED = 1991
# Create a LOOONNNGGG dataframe with each turn of talk,
# it's position in the transcript, the transcript name
# and the speaker tag.


EXCLUDE_TRANSCRIPTS = [
    "2018_91_3C.docx",
    "01_1920_01_2842293_12c.docx",
    "01_1920_05_028_22c.docx",
    "84-2C.docx",
    "103-2C.docx",
    "2019_96_5C.docx",
    "01_1920_05_061_22c.docx",
    "45-2C.docx",
    "61_c.docx",
    "94_c.docx",
    "2018_13_3C.docx",
    "2019_50_5C.docx",
]
# %%

files = os.listdir(start.SHARED_PATH + "excel transcripts/")
files_xlsx = [f for f in files if f[-4:] == "xlsx"]

# %%
df = pd.DataFrame()

for filename in files_xlsx:
    file_path = start.SHARED_PATH + "excel transcripts/"
    data = pd.read_excel(file_path + filename)
    data["doc"] = filename
    df = df.append(data)

df["turn_count"] = df.groupby("doc").cumcount()
df["speaker_turn_count"] = df.groupby(["doc", "Speaker"]).cumcount()
df["total_turn_count"] = df.reset_index().index

# %%


def randomly_select_coach_turn(df):
    random_df = df.sample()
    random_df = random_df.reset_index()
    while random_df.loc[0]["Speaker"] != "Coach":
        random_df = df.sample()
        random_df = random_df.reset_index()
    while random_df.loc[0]["doc"] in EXCLUDE_TRANSCRIPTS:
        random_df = df.sample()
        random_df = random_df.reset_index()

    return random_df


def select_preceding_turns(df, doc, turn_count, preceeding_turns=5):
    new_df = df[df.doc == doc]
    new_df = new_df[
        (new_df.turn_count > turn_count - preceeding_turns)
        & (new_df.turn_count <= turn_count)
    ]
    return new_df


def randomly_select_coach_preceding_turns(df):
    random_df = randomly_select_coach_turn(df)
    new_df = select_preceding_turns(
        df=df, doc=random_df.loc[0]["doc"], turn_count=random_df.loc[0]["turn_count"]
    )
    return new_df


# random_df = pd.DataFrame()
# for i in range(16):
#     random_df = random_df.append(randomly_select_coach_turn(df))

wb = load_workbook(start.DATA_PATH + "certification1.xlsx")
ws = wb.active

row = 2
for i in range(20):
    random_df = randomly_select_coach_preceding_turns(df)
    random_df = random_df.reset_index()
    for item in random_df.index:
        ws.cell(row=row, column=1).value = random_df.loc[item]["Time-stamp"]
        ws.cell(row=row, column=2).value = random_df.loc[item]["Speaker"]
        ws.cell(row=row, column=3).value = random_df.loc[item]["Text"]

        row = row + 1
    row = row + 2

wb.save(start.DATA_PATH + "certification1.xlsx")


wb = load_workbook(start.DATA_PATH + "certification2.xlsx")
ws = wb.active
row = 2
for i in range(20):
    random_df = randomly_select_coach_preceding_turns(df)
    random_df = random_df.reset_index()
    for item in random_df.index:
        ws.cell(row=row, column=1).value = random_df.loc[item]["Time-stamp"]
        ws.cell(row=row, column=2).value = random_df.loc[item]["Speaker"]
        ws.cell(row=row, column=3).value = random_df.loc[item]["Text"]

        row = row + 1
    row = row + 2
wb.save(start.DATA_PATH + "certification2.xlsx")


wb = load_workbook(start.DATA_PATH + "certification3.xlsx")
ws = wb.active
row = 2
for i in range(20):
    random_df = randomly_select_coach_preceding_turns(df)
    random_df = random_df.reset_index()
    for item in random_df.index:
        ws.cell(row=row, column=1).value = random_df.loc[item]["Time-stamp"]
        ws.cell(row=row, column=2).value = random_df.loc[item]["Speaker"]
        ws.cell(row=row, column=3).value = random_df.loc[item]["Text"]

        row = row + 1
    row = row + 2

wb.save(start.DATA_PATH + "certification3.xlsx")


wb = load_workbook(start.DATA_PATH + "certification4.xlsx")
ws = wb.active
row = 2
for i in range(20):
    random_df = randomly_select_coach_preceding_turns(df)
    random_df = random_df.reset_index()
    for item in random_df.index:
        ws.cell(row=row, column=1).value = random_df.loc[item]["Time-stamp"]
        ws.cell(row=row, column=2).value = random_df.loc[item]["Speaker"]
        ws.cell(row=row, column=3).value = random_df.loc[item]["Text"]

        row = row + 1
    row = row + 2

wb.save(start.DATA_PATH + "certification4.xlsx")


# %%
