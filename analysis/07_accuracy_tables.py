# %%
import pandas as pd
import numpy as np
from openpyxl import load_workbook


import statsmodels.formula.api as smf


from moves.library import start

# %%
df_long = pd.read_csv(start.DATA_PATH + "clean/final_long.csv")
df_wide = pd.read_csv(start.DATA_PATH + "clean/final_wide.csv")


df_agg_moves = (
    df_long[["week", "coder", "context", "precision", "recall", "accuracy", "code"]]
    .groupby(by=["week", "coder", "context"])
    .mean()
)

df_agg_within_week = df_agg_moves.groupby(by=["week", "context"]).mean()


def print_summary(filename: str, statistic: str):

    file = start.RESULTS_PATH + filename
    wb = load_workbook(file)
    ws = wb.active

    start_row = 3

    row_n = start_row
    ins = []
    outs = []
    diffs = []

    for week in [1, 2, 3, 4]:
        in_value = df_agg_within_week.loc[(week, "in"), statistic]
        ins.append(in_value)
        out_value = df_agg_within_week.loc[(week, "out"), statistic]
        outs.append(out_value)
        ws.cell(row=row_n, column=2).value = in_value
        ws.cell(row=row_n, column=3).value = out_value
        diff = in_value - out_value
        diffs.append(diff)
        ws.cell(row=row_n, column=4).value = diff
        row_n = row_n + 1

    ws.cell(row=row_n, column=2).value = sum(ins) / len(ins)
    ws.cell(row=row_n, column=3).value = sum(outs) / len(outs)
    ws.cell(row=row_n, column=4).value = sum(diffs) / len(diffs)

    wb.save(file)


print_summary("accuracy_by_week.xlsx", "accuracy")
print_summary("precision_by_week.xlsx", "precision")
print_summary("recall_by_week.xlsx", "recall")


df_agg = df_summary.groupby(["context"]).mean()

# %% Accuracy

file = start.RESULTS_PATH + "main_results.xlsx"
wb = load_workbook(file)
ws = wb.active


def reg_to_excel(df: pd.DataFrame, outcome: str, start_row: int, start_col: int):
    df = df.rename(columns={outcome: "outcome"})

    x_list = [
        "in_context",
        "C(week)[T.2]",
        "C(week)[T.3]",
        "C(week)[T.4]",
        "C(coder)[T.2]",
        "C(coder)[T.3]",
        "C(coder)[T.4]",
    ]

    formula = "outcome ~ in_context + C(move) + C(week) + C(coder)"
    result = smf.ols(formula, data=df).fit(
        cov_type="cluster", cov_kwds={"groups": df["ID"]}, use_t=False
    )
    print(result.summary())

    col_n = start_col
    row_n = start_row
    for x in x_list:
        p = result.pvalues[x]
        coef = result.params[x].round(3)
        if p >= 0.05:
            coef = str(coef)
        if p < 0.05 and p > 0.01:
            coef = str(coef) + "*"
        if p < 0.01 and p > 0.001:
            coef = str(coef) + "**"
        if p < 0.001:
            coef = str(coef) + "***"
        ws.cell(row=row_n, column=col_n).value = coef
        row_n = row_n + 1
        ws.cell(row=row_n, column=col_n).value = "(" + str(result.bse[x].round(3)) + ")"
        row_n = row_n + 1

        wb.save(file)

    return result.summary()


# %%

reg_to_excel(df=df_long, outcome="accuracy", start_row=3, start_col=2)
reg_to_excel(
    df=df_long.dropna(subset=["precision"]),
    outcome="precision",
    start_row=3,
    start_col=3,
)
reg_to_excel(
    df=df_long.dropna(subset=["recall"]),
    outcome="recall",
    start_row=3,
    start_col=4,
)
reg_to_excel(
    df=df_long.dropna(subset=["code"]),
    outcome="code",
    start_row=3,
    start_col=5,
)

# %%
