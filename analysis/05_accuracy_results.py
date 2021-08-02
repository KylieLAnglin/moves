# %%
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook


import statsmodels.formula.api as smf


from moves.library import start

# %%
df = pd.read_csv(start.DATA_PATH + "clean/final_full.csv")

# %%
coder1 = df[(df.week == 1) | (df.week == 3)][
    ["ID", "incontext_a", "week", "master", "move", "move_label"]
].rename(columns={"incontext_a": "code"})
coder1["context"] = "in"
coder1 = coder1.append(
    df[(df.week == 2) | (df.week == 4)][
        ["ID", "outcontext_a", "week", "master", "move", "move_label"]
    ].rename(columns={"outcontext_a": "code"})
)
coder1["context"] = np.where(coder1.context == "in", "in", "out")
coder1["correct"] = np.where(coder1.code == coder1.master, 1, 0)
coder1["coder"] = 1


coder2 = df[(df.week == 1) | (df.week == 3)][
    ["ID", "outcontext_a", "week", "master", "move", "move_label"]
].rename(columns={"outcontext_a": "code"})
coder2["context"] = "out"
coder2 = coder2.append(
    df[(df.week == 2) | (df.week == 4)][
        ["ID", "incontext_a", "week", "master", "move", "move_label"]
    ].rename(columns={"incontext_a": "code"})
)
coder2["context"] = np.where(coder2.context == "out", "out", "in")
coder2["correct"] = np.where(coder2.code == coder2.master, 1, 0)
coder2["coder"] = 2


coder3 = df[(df.week == 1) | (df.week == 3)][
    ["ID", "incontext_b", "week", "master", "move", "move_label"]
].rename(columns={"incontext_b": "code"})
coder3["context"] = "in"
coder3 = coder3.append(
    df[(df.week == 2) | (df.week == 4)][
        ["ID", "outcontext_b", "week", "master", "move", "move_label"]
    ].rename(columns={"outcontext_b": "code"})
)
coder3["context"] = np.where(coder3.context == "in", "in", "out")
coder3["correct"] = np.where(coder3.code == coder3.master, 1, 0)
coder3["coder"] = 3


coder4 = df[(df.week == 1) | (df.week == 3)][
    ["ID", "outcontext_b", "week", "master", "move", "move_label"]
].rename(columns={"outcontext_b": "code"})
coder4["context"] = "out"
coder4 = coder4.append(
    df[(df.week == 2) | (df.week == 4)][
        ["ID", "incontext_b", "week", "master", "move", "move_label"]
    ].rename(columns={"incontext_b": "code"})
)
coder4["context"] = np.where(coder4.context == "out", "out", "in")
coder4["correct"] = np.where(coder4.code == coder4.master, 1, 0)
coder4["coder"] = 4

df_full = pd.concat([coder1, coder2, coder3, coder4])[
    ["ID", "week", "coder", "context", "move", "code", "master"]
]
df_full["precision"] = np.where(df_full.code == 1, 0, np.nan)
df_full["precision"] = np.where(
    (df_full.master == 1) & (df_full.code == 1), 1, df_full.precision
)

df_full["recall"] = np.where(df_full.master == 1, 0, np.nan)
df_full["recall"] = np.where(
    (df_full.master == 1) & (df_full.code == 1), 1, df_full.recall
)

df_full["accuracy"] = np.where(df_full.code == df_full.master, 1, 0)

df_summary = (
    df_full[["week", "coder", "context", "precision", "recall", "accuracy", "code"]]
    .groupby(by=["week", "coder", "context"])
    .mean()
)

summary = df_summary.groupby(["context"]).mean()
# There is no real difference in recall, out of context has slightly gigher.
# But there is a difference in precision. Out-of-context more likely to select yes.
# %%
plt.style.use("seaborn")


def myplot(outcome: str, ylabel: str, saveas: str, ymin, ymax):
    fig, ax = plt.subplots()
    ax.plot(
        [1, 2, 3, 4],
        df_summary.xs(1, level=1, drop_level=False)[outcome],
        color="black",
        linestyle="solid",
    )
    ax.plot(
        [1, 2, 3, 4],
        df_summary.xs(3, level=1, drop_level=False)[outcome],
        color="black",
        linestyle="dashed",
    )
    ax.plot(
        [2, 3, 4, 5],
        df_summary.xs(2, level=1, drop_level=False)[outcome],
        color="gray",
        linestyle="solid",
    )
    ax.plot(
        [2, 3, 4, 5],
        df_summary.xs(4, level=1, drop_level=False)[outcome],
        color="gray",
        linestyle="dashed",
    )

    plt.xlabel("Context")
    plt.ylabel(ylabel)
    plt.ylim(ymin, ymax)
    plt.xticks([1, 2, 3, 4, 5], labels=["In", "Out", "In", "Out", "In"])
    plt.savefig(start.RESULTS_PATH + saveas)


myplot("accuracy", "Accuracy", saveas="single_case_accuracy.png", ymin=0.9, ymax=1)
myplot("precision", "Precision", saveas="single_case_precision.png", ymin=0.5, ymax=1)
myplot("recall", "Recall", saveas="single_case_recall.png", ymin=0.5, ymax=1)


# 1:
# My preference was the out-of-context transcripts. Overall, it was easier to follow because I only had to focus on one move during the entire time. I felt as if I could get them done quicker than in context transcripts for the aforementioned reason. I thought the provided context was just enough to make an informed decision.

# 2:
# I thought I would prefer coding the out-of-context transcripts more because it was easy to just mark a 0 or 1 and then move on, but I preferred the in-context transcripts more. We only had eight options, so it was easy for me to remember what I should be looking for. The context also helped me understand the tone of the coaching conversation more, and it was helpful to see the progression of the coach's goals for the session.
# Coder 2 showed the most consistent pattern, which matched her preference.

# 3:
# My preference was definitely out-of-context transcripts. I think this was because if I knew that for that one transcript all I needed to focus was one thing (for example, positive encouragement), I felt like it was easier on my brain to focus solely on positive encouragement turns of talk, and I could find that one part in the turn of talk to identify. When there is a lot to look at, or multiple codes, sometimes it becomes overwhelming for me to be able to identify all of the codes in the turn of talk, or I forget some.

# 4:
# I preferred out of context transcripts as I felt it was easier to tell whether a code was present or not rather than trying to figure out which codes were present (like for in-context).

# %% Coder aggrement

incontext = df[["ID", "week", "incontext_a", "incontext_b"]]
incontext["agree"] = np.where(incontext.incontext_a == incontext.incontext_b, 1, 0)

incontext_summary = incontext[["week", "agree"]].groupby(by=["week"]).mean()

outcontext = df[["ID", "week", "outcontext_a", "outcontext_b"]]
outcontext["agree"] = np.where(outcontext.outcontext_a == outcontext.outcontext_b, 1, 0)

outcontext_summary = outcontext[["week", "agree"]].groupby(by=["week"]).mean()

fig, ax = plt.subplots()
ax.plot(
    [1, 2, 3, 4],
    incontext_summary.agree,
    color="black",
    linestyle="solid",
)
ax.plot(
    [2, 3, 4, 5],
    outcontext_summary.agree,
    color="gray",
    linestyle="solid",
)

plt.xlabel("Context")
plt.ylabel("Coder Agreement")
plt.ylim(0.9, 1)
plt.xticks([1, 2, 3, 4, 5], labels=["In", "Out", "In", "Out", "In"])
plt.savefig(start.RESULTS_PATH + "single_case_agreement")

# %% Accuracy
df_full["in_context"] = np.where(df_full.context == "in", 1, 0)

file = "/Users/kylie/Dropbox/Active/moves/results/main_results.xlsx"
wb = load_workbook(file)
ws = wb.active


def reg_to_excel(df: pd.DataFrame, outcome: str, start_row: int, start_col: int):
    df = df.rename(columns={outcome: "outcome"})

    x_list = [
        "in_context",
        "C(week)[T.2]",
        "C(week)[T.3]",
        "C(week)[T.4]",
        "C(coder)[T.2]",
        "C(coder)[T.3]",
        "C(coder)[T.4]",
    ]

    formula = "outcome ~ in_context + C(week) + C(coder) + C(move)"
    result = smf.ols(formula, data=df).fit(
        cov_type="cluster", cov_kwds={"groups": df["ID"]}, use_t=False
    )

    col_n = start_col
    row_n = start_row
    for x in x_list:
        p = result.pvalues[x]
        coef = result.params[x].round(3)
        if p >= 0.05:
            coef = str(coef)
        if p < 0.05 and p > 0.01:
            coef = str(coef) + "*"
        if p < 0.01 and p > 0.001:
            coef = str(coef) + "**"
        if p < 0.001:
            coef = str(coef) + "***"
        ws.cell(row=row_n, column=col_n).value = coef
        row_n = row_n + 1
        ws.cell(row=row_n, column=col_n).value = "(" + str(result.bse[x].round(3)) + ")"
        row_n = row_n + 1

        wb.save(file)

    return result.summary()


# %%

reg_to_excel(df=df_full, outcome="accuracy", start_row=3, start_col=2)
reg_to_excel(
    df=df_full.dropna(subset=["precision"]),
    outcome="precision",
    start_row=3,
    start_col=3,
)
reg_to_excel(
    df=df_full.dropna(subset=["recall"]),
    outcome="recall",
    start_row=3,
    start_col=4,
)
reg_to_excel(
    df=df_full.dropna(subset=["code"]),
    outcome="code",
    start_row=3,
    start_col=5,
)

# %%
