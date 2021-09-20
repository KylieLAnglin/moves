# %%
import pandas as pd
import numpy as np
from openpyxl import load_workbook


import statsmodels.formula.api as smf


from moves.library import start

df_long = pd.read_csv(start.DATA_PATH + "clean/final_long.csv")

# %%
file = start.RESULTS_PATH + "main_results.xlsx"
wb = load_workbook(file)
ws = wb.active


def reg_to_excel(df: pd.DataFrame, outcome: str, start_row: int, start_col: int):
    df = df.rename(columns={outcome: "outcome"})

    x_list = [
        "simple_scheme",
        "C(coder)[T.2]",
        "C(coder)[T.3]",
        "C(coder)[T.4]",
        "C(move)[T.2.0]",
        "C(move)[T.3.0]",
        "C(move)[T.4.0]",
        "C(move)[T.5.0]",
        "C(move)[T.6.0]",
        "C(move)[T.7.0]",
        "C(move)[T.8.0]",
    ]

    formula = "outcome ~ simple_scheme + C(move) + week + C(coder)"
    result = smf.ols(formula, data=df).fit(
        cov_type="cluster", cov_kwds={"groups": df["ID"]}, use_t=False
    )
    print(result.summary())

    col_n = start_col
    row_n = start_row
    for x in x_list:
        p = result.pvalues[x]
        coef = result.params[x].round(3)
        if p >= 0.05:
            coef = str(coef)
        if p < 0.05 and p > 0.01:
            coef = str(coef) + "*"
        if p < 0.01 and p > 0.001:
            coef = str(coef) + "**"
        if p < 0.001:
            coef = str(coef) + "***"
        ws.cell(row=row_n, column=col_n).value = coef
        row_n = row_n + 1
        ws.cell(row=row_n, column=col_n).value = "(" + str(result.bse[x].round(3)) + ")"
        row_n = row_n + 1

        wb.save(file)

    return result.summary()


# %%

reg_to_excel(df=df_long, outcome="accuracy", start_row=3, start_col=2)
reg_to_excel(
    df=df_long.dropna(subset=["precision"]),
    outcome="precision",
    start_row=3,
    start_col=3,
)
reg_to_excel(
    df=df_long.dropna(subset=["recall"]),
    outcome="recall",
    start_row=3,
    start_col=4,
)


# %% Accuracy
df_long["simple_scheme"] = np.where(df_long.in_context == 0, 1, 0)

df_long["simple_coder1"] = np.where(
    (df_long.simple_scheme == 1) & (df_long.coder == 1), 1, 0
)

df_long["simple_coder2"] = np.where(
    (df_long.simple_scheme == 1) & (df_long.coder == 2), 1, 0
)

df_long["simple_coder3"] = np.where(
    (df_long.simple_scheme == 1) & (df_long.coder == 3), 1, 0
)

df_long["simple_coder4"] = np.where(
    (df_long.simple_scheme == 1) & (df_long.coder == 4), 1, 0
)

# %%
formula = "recall ~ -1 + simple_coder1 + simple_coder2 + simple_coder3 + simple_coder4 + C(coder) + C(move) + week"
# formula = "recall ~ simple_scheme + C(move) + week + C(coder)"
result = smf.ols(formula, data=df_long).fit()
print(result.summary())

outcome = "precision"
df = df_long[df_long.coder == 3]
formula = outcome + " ~ simple_scheme + C(move) + week"
result = smf.ols(formula, data=df.dropna(subset=[outcome])).fit(
    cov_type="cluster",
    cov_kwds={"groups": df.dropna(subset=[outcome])["ID"]},
    use_t=False,
)
print(result.summary())

formula = "accuracy ~ + simple_scheme + week +C(move) + C(coder)"
result = smf.ols(formula, data=df_long).fit(
    cov_type="cluster", cov_kwds={"groups": df_long["ID"]}, use_t=False
)
print(result.summary())
# %%


# %%
