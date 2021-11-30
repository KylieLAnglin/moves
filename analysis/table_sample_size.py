# %%
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook


import statsmodels.formula.api as smf


from moves.library import start

# %%
df = pd.read_csv(start.CC_PATH + "data/clean/final_wide.csv")

# %% Sample Descriptives

file = start.CC_PATH + "results/" + "sample_descriptives.xlsx"
wb = load_workbook(file)
ws = wb.active

row = 2
col = 2


ws.cell(row=row, column=col).value = df.doc.nunique()
for col_n, week in zip([3, 4, 5, 6], [1, 2, 3, 4]):
    ws.cell(row=row, column=col_n).value = df[df.week == week].doc.nunique()
row = row + 1
ws.cell(row=row, column=col).value = df.ID.nunique()
for col_n, week in zip([3, 4, 5, 6], [1, 2, 3, 4]):
    ws.cell(row=row, column=col_n).value = df[df.week == week].ID.nunique()
row = row + 1
ws.cell(row=row, column=col).value = len(df[(df.master == 1)])
for col_n, week in zip([3, 4, 5, 6], [1, 2, 3, 4]):
    ws.cell(row=row, column=col_n).value = len(df[(df.week == week) & (df.master == 1)])
row = row + 1
for move in [1, 2, 3, 4, 5, 6, 7, 8]:
    ws.cell(row=row, column=col).value = len(df[(df.move == move) & (df.master == 1)])
    for col_n, week in zip([3, 4, 5, 6], [1, 2, 3, 4]):
        ws.cell(row=row, column=col_n).value = len(
            df[(df.move == move) & (df.week == week) & (df.master == 1)]
        )
    row = row + 1


ws.cell(row=row, column=col).value = len(
    df[
        (df.incontext_a == 1)
        & (df.incontext_b == 1)
        & (df.outcontext_a == 1)
        & (df.outcontext_b == 1)
    ]
)
for col_n, week in zip([3, 4, 5, 6], [1, 2, 3, 4]):
    ws.cell(row=row, column=col_n).value = len(
        df[
            (df.week == week)
            & (df.incontext_a == 1)
            & (df.incontext_b == 1)
            & (df.outcontext_a == 1)
            & (df.outcontext_b == 1)
        ]
    )
row = row + 1

df["in_context_correct"] = np.where(
    (df.incontext_a == df.master) & (df.incontext_b == df.master), 1, 0
)
ws.cell(row=row, column=col).value = len(df[(df.in_context_correct == 1)])
for col_n, week in zip([3, 4, 5, 6], [1, 2, 3, 4]):
    ws.cell(row=row, column=col_n).value = len(
        df[(df.week == week) & (df.in_context_correct == 1)]
    )
row = row + 1

df["out_context_correct"] = np.where(
    (df.outcontext_a == df.master) & (df.outcontext_b == df.master), 1, 0
)
ws.cell(row=row, column=col).value = len(df[(df.out_context_correct == 1)])
for col_n, week in zip([3, 4, 5, 6], [1, 2, 3, 4]):
    ws.cell(row=row, column=col_n).value = len(
        df[(df.week == week) & (df.out_context_correct == 1)]
    )
row = row + 1

df["in_context_only"] = np.where(
    (df.in_context_correct == 1) & (df.out_context_correct == 0), 1, 0
)
ws.cell(row=row, column=col).value = len(df[(df.in_context_only == 1)])
for col_n, week in zip([3, 4, 5, 6], [1, 2, 3, 4]):
    ws.cell(row=row, column=col_n).value = len(
        df[(df.week == week) & (df.in_context_only == 1)]
    )
row = row + 1

df["out_context_only"] = np.where(
    (df.in_context_correct == 0) & (df.out_context_correct == 1), 1, 0
)
ws.cell(row=row, column=col).value = len(df[(df.out_context_only == 1)])
for col_n, week in zip([3, 4, 5, 6], [1, 2, 3, 4]):
    ws.cell(row=row, column=col_n).value = len(
        df[(df.week == week) & (df.out_context_only == 1)]
    )
row = row + 1

wb.save(file)


# %%
# number of master codes
# number all agree
# number in-context agree and are correct
# number out-context agree and are correct
# number in and out agree but incontext correct
# number in and out agree but out context correct
# number incontext agree and correct, outcontext disagree
