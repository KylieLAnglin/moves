# %%
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook


import statsmodels.formula.api as smf


from moves.library import start

# %%
df_long = pd.read_csv(start.DATA_PATH + "clean/final_long.csv")
# %%

file = start.RESULTS_PATH + "hte_results.xlsx"
wb = load_workbook(file)
ws = wb.active


def reg_to_excel(df: pd.DataFrame, outcome: str, start_row: int, start_col: int):
    df = df.rename(columns={outcome: "outcome"})

    formula = "outcome ~ in_context + C(week) + C(coder)"

    col_n = start_col
    row_n = start_row

    subgroup_list = [1, 2, 3, 4, 5, 6, 7, 8]

    for subgroup in subgroup_list:
        sub_df = df[df.move == subgroup]
        result = smf.ols(formula, data=sub_df).fit()
        p = result.pvalues["in_context"]
        coef = result.params["in_context"].round(3)
        if p >= 0.05:
            coef = str(coef)
        if p < 0.05 and p > 0.01:
            coef = str(coef) + "*"
        if p < 0.01 and p > 0.001:
            coef = str(coef) + "**"
        if p < 0.001:
            coef = str(coef) + "***"
        ws.cell(row=row_n, column=col_n).value = coef
        row_n = row_n + 1
        ws.cell(row=row_n, column=col_n).value = (
            "(" + str(result.bse["in_context"].round(3)) + ")"
        )
        row_n = row_n + 1

    wb.save(file)

    return result.summary()


reg_to_excel(df=df_long, outcome="accuracy", start_row=3, start_col=2)
reg_to_excel(
    df=df_long.dropna(subset=["precision"]),
    outcome="precision",
    start_row=3,
    start_col=3,
)
reg_to_excel(
    df=df_long.dropna(subset=["recall"]),
    outcome="recall",
    start_row=3,
    start_col=4,
)

subgroup_list = [1, 2, 3, 4, 5, 6, 7, 8]

row = 3
col = 5
for subgroup in subgroup_list:
    sub_df = df[df.move == subgroup]
    sub_df = sub_df[sub_df.master == 1]
    positives = len(sub_df)
    ws.cell(row=row, column=col).value = positives
    row = row + 2

wb.save(file)


# %%

file = start.RESULTS_PATH + "results_hte.xlsx"
wb = load_workbook(file)
ws = wb.active


def reg_to_excel(df: pd.DataFrame, outcome: str, start_row: int, start_col: int):
    df = df.rename(columns={outcome: "outcome"})

    x_list = [
        "in_context",
        "in_context:C(move)[T.2.0]",
        "in_context:C(move)[T.3.0]",
        "in_context:C(move)[T.4.0]",
        "in_context:C(move)[T.5.0]",
        "in_context:C(move)[T.6.0]",
        "in_context:C(move)[T.7.0]",
        "in_context:C(move)[T.8.0]",
    ]

    formula = "outcome ~ in_context * C(move) + C(week) + C(coder)"
    result = smf.ols(formula, data=df).fit(
        cov_type="cluster", cov_kwds={"groups": df["ID"]}, use_t=False
    )
    print(result.summary())

    col_n = start_col
    row_n = start_row
    for x in x_list:
        p = result.pvalues[x]
        coef = result.params[x].round(3)
        if p >= 0.05:
            coef = str(coef)
        if p < 0.05 and p > 0.01:
            coef = str(coef) + "*"
        if p < 0.01 and p > 0.001:
            coef = str(coef) + "**"
        if p < 0.001:
            coef = str(coef) + "***"
        ws.cell(row=row_n, column=col_n).value = coef
        row_n = row_n + 1
        ws.cell(row=row_n, column=col_n).value = "(" + str(result.bse[x].round(3)) + ")"
        row_n = row_n + 1

        wb.save(file)

    return result.summary()


# %%

reg_to_excel(df=df_long, outcome="accuracy", start_row=3, start_col=2)
reg_to_excel(
    df=df_long.dropna(subset=["precision"]),
    outcome="precision",
    start_row=3,
    start_col=3,
)
reg_to_excel(
    df=df_long.dropna(subset=["recall"]),
    outcome="recall",
    start_row=3,
    start_col=4,
)


# %%


# %%
