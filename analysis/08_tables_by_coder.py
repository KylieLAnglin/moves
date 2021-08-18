import pandas as pd
import numpy as np
from openpyxl import load_workbook
import statsmodels.formula.api as smf

from moves.library import start

df_long = pd.read_csv(start.DATA_PATH + "clean/final_long.csv")

# %%
df_agg_moves = (
    df_long[["week", "coder", "context", "precision", "recall", "accuracy", "code"]]
    .groupby(by=["week", "coder", "context"])
    .mean()
)

df_agg_weeks = df_agg_moves.groupby(by=["coder", "context"]).mean()
# %%

df_long["phase"] = np.where(((df_long.week == 1) | (df_long.week == 2)), 1, 2)


def table_by_coder(filename: str, statistic: str, df=pd.DataFrame):

    df = df.dropna(subset=["ID", statistic])

    file = start.RESULTS_PATH + filename
    wb = load_workbook(file)
    ws = wb.active

    row_num = 2
    for coder in [1, 2, 3, 4]:

        formula = statistic + " ~ in_context + C(phase)"
        result = smf.ols(formula, data=df[df.coder == coder]).fit()
        print(result.summary())

        ws.cell(row=row_num, column=2).value = (
            result.params["Intercept"] + result.params["in_context"]
        )
        ws.cell(row=row_num, column=3).value = result.params["Intercept"]
        ws.cell(row=row_num, column=4).value = result.params["in_context"]
        ws.cell(row=row_num, column=5).value = result.pvalues["in_context"]

        row_num = row_num + 1

    wb.save(file)


# %%
table_by_coder(filename="accuracy_by_coder.xlsx", statistic="accuracy", df=df_long)
table_by_coder(filename="precision_by_coder.xlsx", statistic="precision", df=df_long)
table_by_coder(filename="recall_by_coder.xlsx", statistic="recall", df=df_long)

# %%


def reg_to_excel(df: pd.DataFrame, outcome: str, start_row: int, start_col: int):
    df = df.rename(columns={outcome: "outcome"})

    x_list = [
        "in_context",
        "C(week)[T.2]",
        "C(week)[T.3]",
        "C(week)[T.4]",
        "C(coder)[T.2]",
        "C(coder)[T.3]",
        "C(coder)[T.4]",
    ]

    formula = "outcome ~ in_context + C(move) + C(week) + C(coder)"
    result = smf.ols(formula, data=df).fit(
        cov_type="cluster", cov_kwds={"groups": df["ID"]}, use_t=False
    )
    print(result.summary())

    col_n = start_col
    row_n = start_row
    for x in x_list:
        p = result.pvalues[x]
        coef = result.params[x].round(3)
        if p >= 0.05:
            coef = str(coef)
        if p < 0.05 and p > 0.01:
            coef = str(coef) + "*"
        if p < 0.01 and p > 0.001:
            coef = str(coef) + "**"
        if p < 0.001:
            coef = str(coef) + "***"
        ws.cell(row=row_n, column=col_n).value = coef
        row_n = row_n + 1
        ws.cell(row=row_n, column=col_n).value = "(" + str(result.bse[x].round(3)) + ")"
        row_n = row_n + 1

        wb.save(file)

    return result.summary()