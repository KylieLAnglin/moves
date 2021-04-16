# %%
import os
import re
import fnmatch
import collections

import pandas as pd
import docx
from openpyxl import load_workbook

import random
from moves.library import start
from moves.library import process_transcripts

# %%

# %%
def transcript_to_excel(transcript, excel_template: str, excel_file: str):

    wb = load_workbook(excel_template)
    ws = wb.active

    row = 2
    col = 1
    for time in transcript.time_stamps:
        ws.cell(row=row, column=col).value = str(time)
        row = row + 1

    row = 2
    col = 2
    speaker_tags = [tag[0] if tag else "" for tag in transcript.speaker_tags]
    current_speaker = ""
    for speaker in speaker_tags:
        if speaker:
            current_speaker = speaker
        ws.cell(row=row, column=col).value = str(current_speaker)
        row = row + 1

    row = 2
    col = 3
    for text in transcript.text:
        ws.cell(row=row, column=col).value = text
        row = row + 1

    wb.save(excel_file)


# %%

# PATH = "/Users/kylie/Dropbox/1 Projects/Coaching Moves/"
# doc_file = "29_c_Transcript-updated"

# transcript_to_excel(
#     excel_template=PATH + "template.xlsx",
#     doc_file=PATH + doc_file + ".docx",
#     excel_file=PATH + doc_file + ".xlsx",
# )

files = [
    filename
    for filename in os.listdir(start.SHARED_PATH + "uncoded transcripts/")
    if fnmatch.fnmatch(filename, "*.docx") and not filename.startswith("~$")
]

for filename in files:
    # paragraphs = extract_paragraphs(doc_file=start.TRANSCRIPTS_PATH + filename)
    # transcript = extract_data_from_go_transcript(turns_of_talk=paragraphs)
    transcript = process_transcripts.word_to_transcript(
        doc_file=start.TRANSCRIPTS_PATH + filename
    )
    filename = filename[:-5]
    if transcript:
        transcript_to_excel(
            transcript=transcript,
            excel_template=start.SHARED_PATH + "template.xlsx",
            excel_file=start.SHARED_PATH + "excel transcripts/" + filename + ".xlsx",
        )


# %%

# Create random order of files for coding
random.seed(10)
random.shuffle(files)
random_df = pd.DataFrame(files)
random_df.to_csv(start.SHARED_PATH + "random_order.csv")
