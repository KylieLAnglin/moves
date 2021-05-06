# %%
import random

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


from moves.library import start

random_order = pd.read_csv(start.SHARED_PATH + "random_order.csv")

transcript_files = list(random_order.tail(10)["0"])
transcript_files = [f[:-4] + "xlsx" for f in transcript_files]

df = pd.DataFrame()

for filename in transcript_files:
    file_path = start.SHARED_PATH + "excel transcripts/"
    data = pd.read_excel(file_path + filename)
    data["doc"] = filename
    df = df.append(data)

df["turn_count"] = df.groupby("doc").cumcount()
df["speaker_turn_count"] = df.groupby(["doc", "Speaker"]).cumcount()
df["total_turn_count"] = df.reset_index().index

df = df[df.Speaker == "Coach"]
df = df.sample(len(df))

wb = Workbook()
ws = wb.active
ws["A1"] = "TIME SPENT CODING IN SECONDS: "
ws["A2"] = "MOVE: "
ws["B2"] = "TellForward - Suggestion"
ws["A3"] = "Text"
ws["B3"] = "Code"

for r in dataframe_to_rows(df[["Text"]], index=False, header=False):
    ws.append(r)

wb.save(
    start.SHARED_PATH
    + "coding/2021 05 03 Wen Out of Context/1_tell_forward_suggestion.xlsx"
)


# %%
